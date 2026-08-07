"""
excel_parser.py — normalizes inventory_weekly.xlsx (sheet: weekly_counts).

Known issue left untouched here (fixed in cleaning): week_starting is
written in two date formats within the same column.
"""

import os
import openpyxl
from parsers.base import NormalizedRecord, ParseResult


def parse_excel_source(source_name: str, file_path: str, sheet: str = "weekly_counts") -> ParseResult:
    result = ParseResult(source=source_name)

    if not os.path.exists(file_path):
        result.fatal_error = f"file not found: {file_path}"
        return result

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet not in wb.sheetnames:
            result.fatal_error = f"sheet '{sheet}' not found in {file_path} (found: {wb.sheetnames})"
            return result
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        result.fatal_error = f"could not read/parse workbook: {e}"
        return result

    if not rows:
        result.fatal_error = "sheet is empty"
        return result

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]
    result.rows_in = len(data_rows)

    for i, row in enumerate(data_rows):
        try:
            row_dict = dict(zip(header, row))
            week_starting_raw = row_dict.get("week_starting")
            # openpyxl may hand back a datetime object for one format and a
            # plain string for the other, depending on how Excel stored the cell.
            week_starting_str = (
                week_starting_raw.strftime("%Y-%m-%d")
                if hasattr(week_starting_raw, "strftime")
                else str(week_starting_raw) if week_starting_raw is not None else None
            )
            rec = NormalizedRecord(
                source="inventory_weekly",
                record_id=f"{row_dict.get('sku')}-{week_starting_str}-{i}",
                date=week_starting_str,   # raw string, two formats possible — cleaning reconciles
                attrs={
                    "sku": row_dict.get("sku"),
                    "item": row_dict.get("item"),
                    "units_ordered": _to_num(row_dict.get("units_ordered")),
                    "units_sold": _to_num(row_dict.get("units_sold")),
                    "units_wasted": _to_num(row_dict.get("units_wasted")),  # None = not recorded
                    "unit_cost_sar": _to_num(row_dict.get("unit_cost_sar")),
                },
                raw=row_dict,
            )
            result.records.append(rec)
        except Exception as e:
            result.errors.append(f"row {i}: {e}")

    result.rows_out = len(result.records)
    return result


def _to_num(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None
